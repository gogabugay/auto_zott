import openpyxl
import datetime
from datetime import date, timedelta
from dateutil import parser
otchet = openpyxl.load_workbook('Zott_otchet.xlsx')
moscow = openpyxl.load_workbook('Zott.xlsx')
saint_p = openpyxl.load_workbook('Zott_spb.xlsx')

victoria = otchet['Виктория']
victoria_m = moscow['Виктория']
result = otchet['Отчёт']

today = date.today()
friday = date.today() - timedelta(3)
result.cell(row = 2, column = 3, value = today)
#print(yesterday.strftime('%d.%m.%y'))
#print(victoria.cell(row=26,column=13).value)

for i in range (2,26):
    victoria.cell(row=i, column = 13, value = victoria_m.cell(row=i+1, column=5).value.date())
#lenta.cell(row=65, column=30, value='=COUNTIF(O3:O62'+',"='+ yesterday.strftime('%d.%m.%y')+'"')

for k in range (2,26):
    for m in range(15,27):
        victoria.cell(row=k, column=m, value=victoria_m.cell(row=k+1,column=m-8).value)


lenta = otchet['Лента']
lenta_m = moscow['Лента']

for j in range (2,11):
    lenta.cell(row=j, column=13, value = lenta_m.cell(row=j, column=5).value.date())

for l in range (2,11):
    for a in range (15,23):
        lenta.cell(row=l, column=a, value = lenta_m.cell(row=l, column=a-8).value)

globus = otchet['ГиперГлобус']
globus_m = moscow['ГиперГлобус']

for b in range (2,9):
    globus.cell(row=b, column=13, value=globus_m.cell(row=b, column=5).value.date())

for c in range (2,9):
    for i in range (15,27):
        globus.cell(row=c, column=i, value=globus_m.cell(row=c, column=i-8).value)

karusel = otchet['Карусель']
karusel_m = moscow['Карусель']

for b in range (2,22):
    karusel.cell(row=b, column=12, value=karusel_m.cell(row=b, column=5).value.date())

for c in range (2,22):
    for i in range (14,26):
        karusel.cell(row=c, column=i, value=karusel_m.cell(row=c, column=i-7).value)

metro = otchet['Метро']
metro_m = moscow['Метро']

for b in range (2,20):
    metro.cell(row=b, column=13, value=metro_m.cell(row=b+1, column=5).value.date())

for c in range (2,20):
    for i in range (15,34):
        metro.cell(row=c, column=i, value=metro_m.cell(row=c+1, column=i-8).value)

perek = otchet['Перекрёсток']
perek_m = moscow['Перекрёсток']

for b in range (2,89):
    perek.cell(row=b, column=13, value=perek_m.cell(row=b, column=6).value.date())

for c in range (2,89):
    for i in range (15,21):
        perek.cell(row=c, column=i, value=perek_m.cell(row=c, column=i-8).value)

okay = otchet['Окей']
okay_m = moscow['Окей']

for b in range (2,12):
    okay.cell(row=b, column=13, value=okay_m.cell(row=b, column=5).value.date())

for c in range (2,12):
    for i in range (15,21):
        okay.cell(row=c, column=i, value=okay_m.cell(row=c, column=i-8).value)

alie_parusa = otchet['Алые паруса']
alie_parusa_m = moscow['Алые паруса']

alie_parusa.cell(row=2, column=10, value=okay_m.cell(row=2, column=5).value.date())

for i in range (12,28):
    alie_parusa.cell(row=2, column=i, value=alie_parusa_m.cell(row=c, column=i-5).value)


lentaspb = otchet['Лента СПб']
lenta_s = saint_p['Лента СПб']

for i in range(2,30):
    lentaspb.cell(row=i, column = 13, value = friday)

for c in range (2,30):
    for i in range (15,22):
        lentaspb.cell(row=c, column=i, value=lenta_s.cell(row=c, column=i-1).value)

karuselspb = otchet['Карусель СПб']
karusel_s = saint_p['Карусель СПб']

for i in range(2,16):
    karuselspb.cell(row=i, column = 13, value = friday)

for c in range (2,16):
    for i in range (15,26):
        karuselspb.cell(row=c, column=i, value=karusel_s.cell(row=c, column=i-1).value)

metrospb = otchet['Метро СПб']
metro_s = saint_p['Метро СПб']

for i in range(2,5):
    metrospb.cell(row=i, column = 13, value = friday)

for c in range (2,5):
    for i in range (15,31):
        metrospb.cell(row=c, column=i, value=metro_s.cell(row=c, column=i-1).value)

'''lime = otchet['Лайм СПб']
lime_s = saint_p['Лайм СПб']

for i in range(2,13):
    lime.cell(row=i, column = 13, value = friday)

for c in range (2,13):
    for i in range (15,25):
        lime.cell(row=c, column=i, value=lime_s.cell(row=c, column=i-1).value)'''

spar = otchet['Спар СПб']
spar_s = saint_p['Спар СПб']

for i in range(2,18):
    spar.cell(row=i, column = 13, value = friday)

for c in range (2,18):
    for i in range (15,23):
        spar.cell(row=c, column=i, value=spar_s.cell(row=c, column=i).value)

okayspb = otchet['Окей СПб']
okay_s = saint_p['ОКЕЙ СПб']

for i in range(2,23):
    okayspb.cell(row=i, column = 13, value = friday)

for c in range (2,23):
    for i in range (15,20):
        okayspb.cell(row=c, column=i, value=okay_s.cell(row=c, column=i+4).value)

auchan_s = otchet ['Ашан СПБ']
auchan_spb = saint_p ['Ашан СПб']

for i in range (2,11):
    auchan_s.cell(row=i, column=12, value=friday)

for c in range (2,11):
    for i in range (14,29):
        auchan_s.cell(row=c, column=i, value=auchan_spb.cell(row=c, column=i+2).value)

auchan = otchet['Ашан Регион']
auchan_m = moscow['Ашан Регион']
print (auchan_m.cell(row=2, column=6).value)

for b in range (2,10):
    auchan.cell(row=b, column=13, value=auchan_m.cell(row=b, column=6).value.date())

for c in range (2,10):
    for i in range (15,26):
        auchan.cell(row=c, column=i, value=auchan_m.cell(row=c, column=i-7).value)


lentaregion = otchet['Лента Регион']
lenta_reg = moscow['Лента Регион']

for b in range (2,21):
    lentaregion.cell(row=b, column=13, value=lenta_reg.cell(row=b, column=6).value.date())

for c in range (2,21):
    for i in range (15,22):
        lentaregion.cell(row=c, column=i, value=lenta_reg.cell(row=c, column=i-7).value)

metroregion = otchet['Метро Регион']
metro_reg = moscow['Метро Регион']

for b in range (2,9):
    metroregion.cell(row=b, column=13, value=metro_reg.cell(row=b, column=7).value.date())

for c in range (2,9):
    for i in range (15,34):
        metroregion.cell(row=c, column=i, value=metro_reg.cell(row=c, column=i-6).value)

#копируем комменты
for i in range (2,28):
    victoria.cell(row=i, column=29, value=victoria_m.cell(row=i+1,column=20).value)

for i in range (2,11):
    lenta.cell(row=i, column=24, value=lenta_m.cell(row=i,column=17).value)

for i in range (2,8):
    globus.cell(row=i, column=29, value=globus_m.cell(row=i,column=20).value)

for i in range (2,23):
    karusel.cell(row=i, column=28, value=karusel_m.cell(row=i,column=21).value)

for i in range (2,20):
    metro.cell(row=i, column=34, value=metro_m.cell(row=i+1,column=28).value)

for i in range (2,92):
    perek.cell(row=i, column=22, value=perek_m.cell(row=i,column=15).value)

for i in range (2,11):
    okay.cell(row=i, column=22, value=okay_m.cell(row=i,column=15).value)

alie_parusa.cell(row=2, column=30, value=alie_parusa_m.cell(row=2, column=25).value)

for i in range (2,10):
    auchan.cell(row=i, column=26, value=auchan_m.cell(row=i,column=19).value)

for i in range (2,21):
    lentaregion.cell(row=i, column=23, value=lenta_reg.cell(row=i,column=18).value)

for i in range (2,9):
    metroregion.cell(row=i, column=34, value=metro_reg.cell(row=i,column=28).value)



otchet.save('test' + str(today) + '.xlsx')
#print(lenta.cell(row=38, column=15).value.date())

#test_parser = parser.parse(lenta.cell(row=38, column=15).value, dayfirst=True)
#print(test_parser)
#lenta.cell(row=7, column =15, value = str(lenta.cell(row=7, column=15).value[1:]))

#for l in range (3,30):
#    print(lenta_m.cell(row=l, column=7).value.strftime("%d.%m.%y"))

#print(type(lenta_m.cell(row=3, column=7).value))




#dates_lenta_m=[]
#for l in range (2,30):
    #dates_lenta_m+= datetime.strptime(lenta_m.cell(row=l, column=7).value.strftime("%d.%m.%Y"),'%d.%m.%Y')
    #print(dates_lenta_m)
