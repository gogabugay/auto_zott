import openpyxl
import datetime
from datetime import date, timedelta
from dateutil import parser
otchet = openpyxl.load_workbook('Zott - 14.05.xlsx')
moscow = openpyxl.load_workbook('Zott.xlsx')
saint_p = openpyxl.load_workbook('07.05-13.05.18_ZOTT_ОТЧЕТ _СПб.xlsx')
otchet_1 = openpyxl.load_workbook('Zott - 14.05 (копия).xlsx', data_only=True)


victoria_1 = otchet_1['Виктория']
lenta_1 = otchet_1['Лента']
globus_1 = otchet_1['ГиперГлобус']
karusel_1 = otchet_1['Карусель']
metro_1 = otchet_1['Метро']
perek_1 = otchet_1['Перекрёсток']
okay_1 = otchet_1['Окей']
lenta_s1 = otchet_1['Лента СПб']
karusel_s1 = otchet_1['Карусель СПб']
metro_s1 = otchet_1['Метро СПб']
lime_s1 = otchet_1['Лайм СПб']
spar_s1 = otchet_1['Спар СПб']
okay_s1 = otchet_1['Окей СПб']
auchan_reg1 = otchet_1['Ашан Регион']
lenta_reg1 = otchet_1['Лента Регион']
metro_reg1 = otchet_1['Метро Регион']



victoria = otchet['Виктория']
victoria_m = moscow['Виктория']
result = otchet['Отчёт']

today = date.today()
friday = date.today() - timedelta(3)
result.cell(row = 2, column = 3, value = today)
#print(yesterday.strftime('%d.%m.%y'))
#print(victoria.cell(row=26,column=13).value)


for i in range (2,28):
    victoria.cell(row=i, column = 13, value = victoria_m.cell(row=i, column=5).value.date())
#lenta.cell(row=65, column=30, value='=COUNTIF(O3:O62'+',"='+ yesterday.strftime('%d.%m.%y')+'"')

for k in range (2,28):
    for m in range(15,26):
        victoria.cell(row=k, column=m, value=victoria_m.cell(row=k,column=m-8).value)
        victoria_1.cell(row=k, column=m, value=victoria_m.cell(row=k,column=m-8).value)



#проверка комментов

lenta = otchet['Лента']
lenta_m = moscow['Лента']

for j in range (2,11):
    lenta.cell(row=j, column=13, value = lenta_m.cell(row=j, column=5).value.date())

for l in range (2,11):
    for a in range (15,22):
        lenta.cell(row=l, column=a, value = lenta_m.cell(row=l, column=a-8).value)
        lenta_1.cell(row=l, column=a, value = lenta_m.cell(row=l, column=a-8).value)

globus = otchet['ГиперГлобус']
globus_m = moscow['ГиперГлобус']

for b in range (2,8):
    globus.cell(row=b, column=13, value=globus_m.cell(row=b, column=5).value.date())

for c in range (2,8):
    for i in range (15,22):
        globus.cell(row=c, column=i, value=globus_m.cell(row=c, column=i-8).value)
        globus_1.cell(row=c, column=i, value=globus_m.cell(row=c, column=i-8).value)


karusel = otchet['Карусель']
karusel_m = moscow['Карусель']

for b in range (2,23):
    karusel.cell(row=b, column=12, value=karusel_m.cell(row=b, column=5).value.date())

for c in range (2,23):
    for i in range (14,26):
        karusel.cell(row=c, column=i, value=karusel_m.cell(row=c, column=i-7).value)
        karusel_1.cell(row=c, column=i, value=karusel_m.cell(row=c, column=i-7).value)


metro = otchet['Метро']
metro_m = moscow['Метро']

for b in range (2,20):
    metro.cell(row=b, column=13, value=metro_m.cell(row=b, column=5).value.date())

for c in range (2,20):
    for i in range (15,31):
        metro.cell(row=c, column=i, value=metro_m.cell(row=c, column=i-8).value)
        metro_1.cell(row=c, column=i, value=metro_m.cell(row=c, column=i-8).value)


perek = otchet['Перекрёсток']
perek_m = moscow['Перекрёсток']

for b in range (2,80):
    perek.cell(row=b, column=13, value=perek_m.cell(row=b, column=6).value.date())

for c in range (2,80):
    for i in range (15,21):
        perek.cell(row=c, column=i, value=perek_m.cell(row=c, column=i-8).value)
        perek_1.cell(row=c, column=i, value=perek_m.cell(row=c, column=i-8).value)

okay = otchet['Окей']
okay_m = moscow['Окей']

for b in range (2,11):
    okay.cell(row=b, column=13, value=okay_m.cell(row=b, column=5).value.date())

for c in range (2,11):
    for i in range (15,20):
        okay.cell(row=c, column=i, value=okay_m.cell(row=c, column=i-8).value)
        okay_1.cell(row=c, column=i, value=okay_m.cell(row=c, column=i-8).value)


lentaspb = otchet['Лента СПб']
lenta_s = saint_p['Лента СПб']

for i in range(2,30):
    lentaspb.cell(row=i, column = 13, value = friday)

for c in range (2,30):
    for i in range (15,22):
        lentaspb.cell(row=c, column=i, value=lenta_s.cell(row=c, column=i-1).value)

karuselspb = otchet['Карусель СПб']
karusel_s = saint_p['Карусель СПб']

for i in range(2,16):
    karuselspb.cell(row=i, column = 13, value = friday)

for c in range (2,16):
    for i in range (15,26):
        karuselspb.cell(row=c, column=i, value=karusel_s.cell(row=c, column=i-1).value)

metrospb = otchet['Метро СПб']
metro_s = saint_p['Метро СПб']

for i in range(2,5):
    metrospb.cell(row=i, column = 13, value = friday)

for c in range (2,5):
    for i in range (15,31):
        metrospb.cell(row=c, column=i, value=metro_s.cell(row=c, column=i-1).value)

lime = otchet['Лайм СПб']
lime_s = saint_p['Лайм СПб']

for i in range(2,13):
    lime.cell(row=i, column = 13, value = friday)

for c in range (2,13):
    for i in range (15,25):
        lime.cell(row=c, column=i, value=lime_s.cell(row=c, column=i-1).value)

spar = otchet['Спар СПб']
spar_s = saint_p['Спар СПб']

for i in range(2,18):
    spar.cell(row=i, column = 13, value = friday)

for c in range (2,18):
    for i in range (15,22):
        spar.cell(row=c, column=i, value=spar_s.cell(row=c, column=i-1).value)

okayspb = otchet['Окей СПб']
okay_s = saint_p['ОКЕЙ']

for i in range(2,23):
    okayspb.cell(row=i, column = 13, value = friday)

for c in range (2,23):
    for i in range (15,20):
        okayspb.cell(row=c, column=i, value=okay_s.cell(row=c, column=i+4).value)

auchan = otchet['Ашан Регион']
auchan_m = moscow['Ашан Регион']
#print (auchan_m.cell(row=2, column=6).value)

for b in range (2,10):
    auchan.cell(row=b, column=13, value=auchan_m.cell(row=b, column=6).value.date())

for c in range (2,10):
    for i in range (15,24):
        auchan.cell(row=c, column=i, value=auchan_m.cell(row=c, column=i-7).value)


lentaregion = otchet['Лента Регион']
lenta_reg = moscow['Лента Регион']

for b in range (2,21):
    lentaregion.cell(row=b, column=13, value=lenta_reg.cell(row=b, column=6).value.date())

for c in range (2,21):
    for i in range (15,22):
        lentaregion.cell(row=c, column=i, value=lenta_reg.cell(row=c, column=i-7).value)

metroregion = otchet['Метро Регион']
metro_reg = moscow['Метро Регион']

for b in range (2,9):
    metroregion.cell(row=b, column=13, value=metro_reg.cell(row=b, column=7).value.date())

for c in range (2,9):
    for i in range (15,31):
        metroregion.cell(row=c, column=i, value=metro_reg.cell(row=c, column=i-6).value)

for i in range (2,28):
    count=0
    count1=0
    for j in range (15,26):
        if victoria_1.cell(row=i, column=j).value == 'х':
            count+=1
            victoria_1.cell(row=i, column=14, value=11-count)
        elif victoria_1.cell(row=i, column=j).value==1:
            count1+=1
            victoria_1.cell(row=i, column=26, value=count1)

for g in range (2,28):
    victoria_1.cell(row=g, column=27, value= victoria_1.cell(row=g, column=26).value / victoria_1.cell(row=g, column=14).value)


for i in range (2,28):
    victoria.cell(row=i, column=28, value= victoria_m.cell(row=i,column=20).value)

for j in range(2,28):
    if victoria.cell(row=j, column=28).value==None and victoria_1.cell(row=j, column=27).value<1:
        victoria.cell(row=j, column=28).value = 'Ожидается поставка с' + ' ' + str(today)

for i in range (2,11):
    count=0
    count1=0
    for j in range (15,22):
        if lenta_1.cell(row=i, column=j).value == 'х':
            count+=1
            lenta_1.cell(row=i, column=14, value=7-count)
        elif lenta_1.cell(row=i, column=j).value==1:
            count1+=1
            lenta_1.cell(row=i, column=22, value=count1)

for g in range (2,11):
    lenta_1.cell(row=g, column=23, value= lenta_1.cell(row=g, column=22).value / lenta_1.cell(row=g, column=14).value)


for i in range (2,11):
    lenta.cell(row=i, column=24, value= lenta_m.cell(row=i,column=16).value)


for j in range (2,11):
    if lenta.cell(row=j, column=24).value==None and lenta_1.cell(row=j, column=23).value<1:
        lenta.cell(row=j, column=24, value='Ожидается поставка с' + ' ' + str(today))

for i in range (2,8):
    count=0
    count1=0
    for j in range (15,22):
        if globus_1.cell(row=i, column=j).value == 'х':
            count+=1
            globus_1.cell(row=i, column=14, value=11-count)
        elif globus_1.cell(row=i, column=j).value==1:
            count1+=1
            globus_1.cell(row=i, column=22, value=count1)

for g in range (2,8):
    globus_1.cell(row=g, column=23, value= globus_1.cell(row=g, column=22).value / globus_1.cell(row=g, column=14).value)

for i in range (2,11):
    globus.cell(row=i, column=24, value= lenta_m.cell(row=i,column=16).value)

for i in range (2,8):
    if globus.cell(row=i, column=24).value==None and globus_1.cell(row=i, column=23).value<1:
        globus.cell(row=j, column=24, value='Ожидается поставка с' + ' ' + str(today))

for i in range (2,23):
    count=0
    count1=0
    for j in range (14,26):
        if karusel_1.cell(row=i, column=j).value == 'х':
            count+=1
            karusel_1.cell(row=i, column=13, value=12-count)
        elif karusel_1.cell(row=i, column=j).value==1:
            count1+=1
            karusel_1.cell(row=i, column=26, value=count1)

for g in range (2,23):
    karusel_1.cell(row=g, column=27, value= karusel_1.cell(row=g, column=26).value / karusel_1.cell(row=g, column=13).value)

for i in range (2,23):
    karusel.cell(row=i, column=28, value= lenta_m.cell(row=i,column=21).value)

for i in range (2,23):
    if karusel.cell(row=i, column=28).value==None and karusel_1.cell(row=i, column=27).value<1:
        karusel.cell(row=j, column=28, value='Ожидается поставка с' + ' ' + str(today))

for i in range (2,20):
    count=0
    count1=0
    for j in range (15,31):
        if metro_1.cell(row=i, column=j).value == 'х':
            count+=1
            metro_1.cell(row=i, column=14, value=16-count)
        elif metro_1.cell(row=i, column=j).value==1:
            count1+=1
            metro_1.cell(row=i, column=31, value=count1)

for g in range (2,20):
    metro_1.cell(row=g, column=32, value= metro_1.cell(row=g, column=31).value / metro_1.cell(row=g, column=14).value)

for i in range (2,20):
    metro.cell(row=i, column=33, value= lenta_m.cell(row=i,column=25).value)

for i in range (2,20):
    if metro.cell(row=i, column=33).value==None and metro_1.cell(row=i, column=32).value<1:
        metro.cell(row=j, column=33, value='Ожидается поставка с '+ str(today))

for i in range (2,80):
    count=0
    count1=0
    for j in range (15,21):
        if perek_1.cell(row=i, column=j).value == 'х':
            count+=1
            perek_1.cell(row=i, column=14, value=6-count)
        elif perek_1.cell(row=i, column=j).value==1:
            count1+=1
            perek_1.cell(row=i, column=21, value=count1)

for g in range (2,80):
    perek_1.cell(row=g, column=22, value= perek_1.cell(row=g, column=21).value / perek_1.cell(row=g, column=14).value)

for i in range (2,80):
    perek.cell(row=i, column=23, value= lenta_m.cell(row=i,column=15).value)

for i in range (2,80):
    if perek.cell(row=i, column=23).value==None and perek_1.cell(row=i, column=22).value<1:
        perek.cell(row=j, column=23, value='Ожидается поставка с '+ str(today))

for i in range (2,11):
    count=0
    count1=0
    for j in range (15,20):
        if okay_1.cell(row=i, column=j).value == 'х':
            count+=1
            okay_1.cell(row=i, column=14, value=5-count)
        elif okay_1.cell(row=i, column=j).value==1:
            count1+=1
            okay_1.cell(row=i, column=20, value=count1)

for g in range (2,11):
    okay_1.cell(row=g, column=21, value= okay_1.cell(row=g, column=20).value / okay_1.cell(row=g, column=14).value)

for i in range (2,11):
    okay.cell(row=i, column=22, value= lenta_m.cell(row=i,column=14).value)

#print(okay_1.cell(row=2, column=21).value)
for i in range (2,11):
    if okay.cell(row=i, column=22).value==None and okay_1.cell(row=i, column=21).value<1:
        okay.cell(row=j, column=22, value='Ожидается поставка с '+ str(today))

print((victoria_1.cell(row=24, column=27).value))


#print(lenta.cell(row=38, column=15).value.date())
otchet_1.save('test.xlsx')
otchet.save('zott' + str(today) + '.xlsx')

#test_parser = parser.parse(lenta.cell(row=38, column=15).value, dayfirst=True)
#print(test_parser)
#lenta.cell(row=7, column =15, value = str(lenta.cell(row=7, column=15).value[1:]))

#for l in range (3,30):
#    print(lenta_m.cell(row=l, column=7).value.strftime("%d.%m.%y"))

#print(type(lenta_m.cell(row=3, column=7).value))




#dates_lenta_m=[]
#for l in range (2,30):
    #dates_lenta_m+= datetime.strptime(lenta_m.cell(row=l, column=7).value.strftime("%d.%m.%Y"),'%d.%m.%Y')
    #print(dates_lenta_m)
